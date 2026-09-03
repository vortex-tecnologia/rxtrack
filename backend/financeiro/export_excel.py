# Copyright (c) 2026 Luiz Gustavo. Todos os direitos reservados. Licença Proprietária.
# financeiro/export_excel.py

"""
Exportador de Fechamento de Agregados para arquivo Excel (.xlsx).
Gera uma planilha com abas individuais por motorista e a aba consolidada 'RESULTADO FATURA',
idêntica ao layout e fórmulas da planilha original.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import io


def exportar_fechamento_excel(fechamento):
    """
    Gera o workbook openpyxl com todas as abas e retorna os bytes do arquivo Excel.
    """
    wb = openpyxl.Workbook()
    # Remove sheet padrão inicial
    default_sheet = wb.active

    # Cores e estilos padrão
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subtotal_font = Font(name="Calibri", size=10, bold=True)
    
    title_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=10)
    money_format = '#,##0.00'
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # Identifica todas as bases/UFs presentes no fechamento para colunas dinâmicas
    todas_bases = set()
    linhas_por_motorista = {}
    
    for linha in fechamento.linhas.select_related('motorista', 'manifesto', 'filial_operacao', 'manifesto__veiculo').order_by('motorista__nome_completo', 'data'):
        mot_id = linha.motorista_id
        if mot_id not in linhas_por_motorista:
            linhas_por_motorista[mot_id] = {
                'motorista': linha.motorista,
                'linhas': []
            }
        linhas_por_motorista[mot_id]['linhas'].append(linha)
        for base in (linha.breakdown_bases or {}).keys():
            if base and base != 'OUTROS':
                todas_bases.add(base.upper().strip())

    lista_bases = sorted(list(todas_bases)) if todas_bases else ['RJ', 'SP', 'DF', 'GO']

    # Resumos indexados por motorista
    resumos_dict = {r.motorista_id: r for r in fechamento.resumos.all()}

    # 1. CRIA ABAS INDIVIDUAIS POR MOTORISTA
    for mot_id, dados in linhas_por_motorista.items():
        motorista = dados['motorista']
        linhas = dados['linhas']
        resumo = resumos_dict.get(mot_id)

        # Nome da aba (máximo 31 caracteres no Excel)
        nome_aba = (motorista.nome_completo or f"Motorista_{mot_id}")[:31]
        # Remove caracteres inválidos para nome de aba no Excel
        for ch in [':', '\\', '/', '?', '*', '[', ']']:
            nome_aba = nome_aba.replace(ch, '')
        
        ws = wb.create_sheet(title=nome_aba)
        ws.views.sheetView[0].showGridLines = True

        # Placa do veículo mais recente ou vazio
        placa = ""
        for l in linhas:
            if l.manifesto and l.manifesto.veiculo and l.manifesto.veiculo.placa:
                placa = l.manifesto.veiculo.placa
                break

        # Cabeçalho do motorista
        periodo_str = f"{fechamento.periodo_inicio.strftime('%d/%m/%Y')} A {fechamento.periodo_fim.strftime('%d/%m/%Y')}"
        ws['A1'] = f"PERIODO: {periodo_str}"
        ws['A1'].font = title_font
        ws['A2'] = f"MOTORISTA: {motorista.nome_completo.upper()}"
        ws['A2'].font = title_font
        ws['A3'] = f"CARRO/PLACA: {placa}"
        ws['A3'].font = title_font
        ws['A4'] = f"Obs: {resumo.observacao if resumo and resumo.observacao else ''}"
        ws['A4'].font = Font(name="Calibri", size=9, italic=True)

        # Cabeçalhos da Tabela na Linha 5
        headers = [
            "Nº MANIFESTO", "Data", "Valor da Diária", "Valores Extras", "Localidade Extra",
            "Nº CTE", "Nº Ctrc Realizados", "Valor Entregas", "Nº Coletas", "Coletas Válidas",
            "Total Embarques", "Valor Coletas", "TOTAL DO DIA", "OBSERVAÇÃO"
        ]
        
        # Adiciona colunas dinâmicas para cada base (Entregas e Coletas)
        for base in lista_bases:
            headers.append(f"Entregas {base}")
            headers.append(f"Coletas {base}")

        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=h_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Preenchimento das linhas (a partir da linha 6)
        row_num = 6
        for l in linhas:
            ws.cell(row=row_num, column=1, value=l.manifesto.numero_manifesto if l.manifesto else "")
            ws.cell(row=row_num, column=2, value=l.data.strftime('%d/%m/%Y') if l.data else "")
            
            c_diaria = ws.cell(row=row_num, column=3, value=float(l.valor_diaria))
            c_diaria.number_format = money_format
            
            c_extra = ws.cell(row=row_num, column=4, value=float(l.valor_extra))
            c_extra.number_format = money_format
            
            ws.cell(row=row_num, column=5, value=l.localidade_extra or "")
            ws.cell(row=row_num, column=6, value=l.qtd_ctes)
            ws.cell(row=row_num, column=7, value=l.qtd_ctes_realizados)
            
            c_ventregas = ws.cell(row=row_num, column=8, value=float(l.valor_entregas))
            c_ventregas.number_format = money_format
            
            ws.cell(row=row_num, column=9, value=l.qtd_coletas)
            ws.cell(row=row_num, column=10, value=l.qtd_coletas_validas)
            ws.cell(row=row_num, column=11, value=l.total_embarques)
            
            c_vcoletas = ws.cell(row=row_num, column=12, value=float(l.valor_coletas))
            c_vcoletas.number_format = money_format
            
            c_total = ws.cell(row=row_num, column=13, value=float(l.total_dia))
            c_total.number_format = money_format
            c_total.font = Font(bold=True)
            
            ws.cell(row=row_num, column=14, value=l.observacao or "")

            # Bases dinâmicas
            bk = l.breakdown_bases or {}
            base_col = 15
            for base in lista_bases:
                base_data = bk.get(base, {})
                ws.cell(row=row_num, column=base_col, value=base_data.get('entregas', 0))
                ws.cell(row=row_num, column=base_col + 1, value=base_data.get('coletas', 0))
                base_col += 2

            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).border = thin_border
                ws.cell(row=row_num, column=c).font = normal_font

            row_num += 1

        # Linha de Totais
        last_data_row = row_num - 1
        r_tot = row_num + 1

        ws.cell(row=r_tot, column=2, value="TOTAL DIÁRIAS").font = subtotal_font
        cell_td = ws.cell(row=r_tot, column=3, value=f"=SUM(C6:C{last_data_row})")
        cell_td.number_format = money_format
        cell_td.font = subtotal_font

        ws.cell(row=r_tot, column=10, value="TOTAL SERVIÇOS").font = subtotal_font
        cell_ts = ws.cell(row=r_tot, column=11, value=f"=SUM(K6:K{last_data_row})")
        cell_ts.font = subtotal_font

        ws.cell(row=r_tot, column=12, value="TOTAL PARCIAL").font = subtotal_font
        cell_tp = ws.cell(row=r_tot, column=13, value=f"=SUM(M6:M{last_data_row}) + C{r_tot+1}")
        cell_tp.number_format = money_format
        cell_tp.font = subtotal_font

        # Pedágio e Descontos
        pedagio_val = float(resumo.valor_pedagio) if resumo else 0.0
        desconto_val = float(resumo.valor_desconto) if resumo else 0.0

        ws.cell(row=r_tot + 1, column=2, value="TOTAL PEDÁGIO").font = subtotal_font
        c_ped = ws.cell(row=r_tot + 1, column=3, value=pedagio_val)
        c_ped.number_format = money_format
        c_ped.font = subtotal_font

        ws.cell(row=r_tot + 1, column=10, value="DIÁRIA / SERVIÇO").font = subtotal_font
        c_ds = ws.cell(row=r_tot + 1, column=11, value=f"=IFERROR((C{r_tot}+C{r_tot+1})/K{r_tot}, 0)")
        c_ds.number_format = money_format
        c_ds.font = subtotal_font

        ws.cell(row=r_tot + 1, column=12, value="DESCONTOS").font = subtotal_font
        c_desc = ws.cell(row=r_tot + 1, column=13, value=desconto_val)
        c_desc.number_format = money_format
        c_desc.font = subtotal_font

        # Total Final
        ws.cell(row=r_tot + 2, column=12, value="TOTAL FINAL").font = Font(bold=True, size=11, color="9C0006")
        c_tf = ws.cell(row=r_tot + 2, column=13, value=f"=M{r_tot} + M{r_tot+1}")
        c_tf.number_format = money_format
        c_tf.font = Font(bold=True, size=11, color="9C0006")
        c_tf.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Ajusta largura das colunas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len and len(val_str) < 50:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # 2. CRIA ABA CONSOLIDADA 'RESULTADO FATURA'
    ws_fat = wb.create_sheet(title="RESULTADO FATURA")
    ws_fat.views.sheetView[0].showGridLines = True

    headers_fat = ["NOME MOTORISTA", "CPF DO TITULAR", "DADOS BANCÁRIOS / PIX"]
    for base in lista_bases:
        headers_fat.append(f"VALOR {base}")
    headers_fat.append("VALOR TOTAL")

    for col_idx, h_text in enumerate(headers_fat, start=1):
        cell = ws_fat.cell(row=1, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_fat = 2
    from financeiro.models import DadosBancariosAgregado
    dados_banc_dict = {db.motorista_id: db for db in DadosBancariosAgregado.objects.all()}

    for mot_id, dados in linhas_por_motorista.items():
        motorista = dados['motorista']
        resumo = resumos_dict.get(mot_id)
        banco = dados_banc_dict.get(mot_id)

        ws_fat.cell(row=row_fat, column=1, value=motorista.nome_completo)
        ws_fat.cell(row=row_fat, column=2, value=motorista.cpf or "")
        
        info_banco = ""
        if banco:
            partes = []
            if banco.dados_bancarios: partes.append(banco.dados_bancarios)
            if banco.chave_pix: partes.append(f"PIX: {banco.chave_pix}")
            if banco.titular_pagamento: partes.append(f"Titular: {banco.titular_pagamento}")
            info_banco = " - ".join(partes)
        ws_fat.cell(row=row_fat, column=3, value=info_banco)

        # Valores por base do resumo
        bk_resumo = (resumo.breakdown_bases or {}) if resumo else {}
        col_curr = 4
        soma_bases_row = []
        for base in lista_bases:
            v_base = bk_resumo.get(base, {}).get('valor_total', 0.0)
            c = ws_fat.cell(row=row_fat, column=col_curr, value=float(v_base))
            c.number_format = money_format
            soma_bases_row.append(get_column_letter(col_curr) + str(row_fat))
            col_curr += 1

        # Total do motorista
        total_val = float(resumo.total_final) if resumo else 0.0
        c_tot = ws_fat.cell(row=row_fat, column=col_curr, value=total_val)
        c_tot.number_format = money_format
        c_tot.font = Font(bold=True)

        for c in range(1, len(headers_fat) + 1):
            ws_fat.cell(row=row_fat, column=c).border = thin_border

        row_fat += 1

    # Totais da fatura
    last_fat_row = row_fat - 1
    ws_fat.cell(row=row_fat + 1, column=3, value="TOTAL POR UNIDADE DE NEGÓCIOS").font = subtotal_font
    col_curr = 4
    for base in lista_bases:
        col_let = get_column_letter(col_curr)
        c = ws_fat.cell(row=row_fat + 1, column=col_curr, value=f"=SUM({col_let}2:{col_let}{last_fat_row})")
        c.number_format = money_format
        c.font = subtotal_font
        col_curr += 1

    col_let_tot = get_column_letter(col_curr)
    c_tot_geral = ws_fat.cell(row=row_fat + 1, column=col_curr, value=f"=SUM({col_let_tot}2:{col_let_tot}{last_fat_row})")
    c_tot_geral.number_format = money_format
    c_tot_geral.font = Font(bold=True, size=11, color="1F4E79")

    # Ajuste de largura na fatura
    for col in ws_fat.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len and len(val_str) < 50:
                max_len = len(val_str)
        ws_fat.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # Remove aba vazia inicial
    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    # Salva em memória e retorna bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
